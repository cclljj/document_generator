#!/usr/bin/env python

import re
import argparse
import openpyxl
import zipfile

SN_pattern = re.compile("^[0-9]+$")
NUM_pattern = re.compile("[0-9]*\.?[0-9]+")

def doc_generate(fname, templatefname, prefix):

	wb = openpyxl.load_workbook(fname)
	sheets = wb.get_sheet_names()
	ws = wb.get_sheet_by_name(sheets[0])

	ColNames = {}
	Current  = 0
	for COL in ws.iter_cols(1, ws.max_column):
		ColNames[COL[0].value] = Current
		Current += 1

	for row_cells in ws.iter_rows(min_row=2):
		if not SN_pattern.match(str(row_cells[0].value)):
			break
		replaceText = {}
		for COL in ColNames:
			if type(row_cells[ColNames[COL]].value) == str:
				replaceText[str(COL)] = row_cells[ColNames[COL]].value
			elif type(row_cells[ColNames[COL]].value) == unicode:
				replaceText[str(COL)] = unicode(row_cells[ColNames[COL]].value)
			else:
				replaceText[str(COL)] = str(row_cells[ColNames[COL]].value)

		output_fname = prefix + "/" + str(row_cells[0].value) + ".docx"
		templateDocx = zipfile.ZipFile(templatefname)
		newDocx = zipfile.ZipFile(output_fname, "w")
		for file in templateDocx.filelist:
			content = templateDocx.read(file)
			for key in replaceText.keys():
				if type(replaceText[key]) == unicode:
					content = content.replace("{{" + key + "}}", replaceText[key].encode('utf-8'))
				else:
					content = content.replace("{{" + key + "}}", replaceText[key])
			newDocx.writestr(file.filename, content)
		templateDocx.close()
		newDocx.close()

if __name__ == "__main__":
	parser = argparse.ArgumentParser(description='Retrieve data of one specific project on a specific date')
	parser.add_argument("-d", dest="DATAFILE", type=str, help="Data file name (.xlsx)", required=True)
	parser.add_argument("-t", dest="TEMPLATE", type=str, help="Output template file name (*.docx)", required=True)
	parser.add_argument("-p", dest="PREFIX", default="./", type=str, help="Output folder (default: ./)")
	args = parser.parse_args()

	doc_generate(args.DATAFILE, args.TEMPLATE, args.PREFIX)

